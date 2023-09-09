import os

from flask import render_template, redirect, request, session, flash, url_for
from openpyxl import load_workbook
import subprocess
import pandas as pd
from app import app, USERNAMES, PASSWORDS


def add_user(username, password):
    DB = load_workbook(filename="db.xlsx")
    Users = DB["Users"]
    new_row = Users.max_row + 1
    Users[f"{USERNAMES}{new_row}"] = username
    Users[f"{PASSWORDS}{new_row}"] = password
    DB.save(filename="db.xlsx")

def read_db() -> pd.DataFrame:
    subprocess.Popen(["libreoffice", "--headless", "--convert-to", "csv", "db.xlsx"], stdout=subprocess.DEVNULL, stderr=subprocess.STDOUT).communicate()
    df = pd.read_csv("db.csv")
    return df


@app.route("/", methods=["POST", "GET"])
def base():
    if not session.get("username"):
        return redirect(url_for("login"))
    else:
        Users = read_db()
        username = session.get("username")
        password = Users.query(f"Username == '{username}'")["Password"].values[0]
        return render_template('index.html', name=username, password=password)


@app.route("/admin", methods=["GET"])
def admin():
    username = session.get("username")
    if username != "admin":
        return redirect("/")
    else:
        return render_template("admin.html", name=username)


@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == "GET":
        return render_template("login.html")
    else:
        username = request.form.get("username")
        password = request.form.get("password")

        Users = read_db()
        
        if username not in Users.Username.values:     
            flash('Please check your login details and try again.', 'danger')
            return redirect(url_for('login'))
        elif password != Users.query(f"Username == '{username}'")["Password"].values[0]:
            flash('Please check your login details and try again.', 'danger')
            return redirect(url_for('login'))

        session["username"] = request.form.get("username")
        return redirect("/")


@app.route('/signup', methods=['GET','POST'])
def signup():
    if request.method=='POST':
        username = request.form.get("username")
        password = request.form.get("password")

        Users = read_db()
        if username in Users.Username.values:     
            flash('Username already exists')
            return redirect(url_for('signup'))
        else:
            add_user(username, password)
            session["username"] = username
            return redirect("/")
    else:
        return render_template('signup.html')


@app.route('/logout', methods=['GET'])
def logout():
    if request.method=='GET':
        username = session.get("username")
        session.pop('username', default=None)
        return redirect("/")


if __name__ == "__main__":
    for f in os.listdir("flask_session"):
        os.remove(os.path.join("flask_session", f))

    app.run(debug=False, host='0.0.0.0', threaded=True)