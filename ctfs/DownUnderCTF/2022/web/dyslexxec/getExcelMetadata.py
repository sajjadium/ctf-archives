import sys
import uuid
import os
import shutil
from lxml import etree
from openpyxl import load_workbook
from zipfile import ZipFile

WORKBOOK = "xl/workbook.xml"

def getMetadata(filename):
    
    properties = []

    try:
        wb = load_workbook(filename)
        for e in wb.properties.__elements__:
            properties.append(
                {
                    "Fieldname" : e, 
                    "Attribute" : None,
                    "Value" : getattr(wb.properties, e)
                }
            )
        for s in wb.sheetnames:
            properties.append(
                {
                    "Fieldname" : "sheet", 
                    "Attribute" : s,
                    "Value" : None
                }
            )
    except Exception:
        print("error loading workbook")
        return None

    return properties
    


def extractWorkbook(filename, outfile="xml"):
    with ZipFile(filename, "r") as zip:
        zip.extract(WORKBOOK, outfile)



def findInternalFilepath(filename):
    try:
        prop = None
        parser = etree.XMLParser(load_dtd=True, resolve_entities=True)
        tree = etree.parse(filename, parser=parser)
        root = tree.getroot()
        internalNode = root.find(".//{http://schemas.microsoft.com/office/spreadsheetml/2010/11/ac}absPath")
        if internalNode != None:
            prop = {
                "Fieldname":"absPath",
                "Attribute":internalNode.attrib["url"],
                "Value":internalNode.text
            }
        return prop

    except Exception:
        print("couldnt extract absPath")
        return None

    



if __name__ == "__main__":
    if len(sys.argv) == 2:
        filename = sys.argv[1]
    else:
        print("Usage:", sys.argv[0], "<filename>")
        exit(1)
    
    tmpFolder = "./uploads/" + str(uuid.uuid4())
    os.mkdir(tmpFolder)

    properties = getMetadata(filename)

    extractWorkbook(filename, tmpFolder)

    workbook = tmpFolder + "/" + WORKBOOK
    properties.append(findInternalFilepath(workbook))

    for p in properties:
        print(p)

    print("Removing tmp folder:", workbook)
    shutil.rmtree(tmpFolder)